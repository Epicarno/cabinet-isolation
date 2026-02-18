# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r"c:\Users\Администратор\Desktop\Modules"

# Read KKS reference
f = os.path.join(base, "Применение KKS - ОП СПб.xlsx")
wb = openpyxl.load_workbook(f, data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*80}")
    print(f"=== Sheet: {sname} ===")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    for r in range(1, min(ws.max_row + 1, 60)):
        vals = []
        for col in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                vals.append((openpyxl.utils.get_column_letter(col), v))
        if vals:
            print(f"  Row {r}: {vals}")

wb.close()
