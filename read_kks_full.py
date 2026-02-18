# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r"c:\Users\Администратор\Desktop\Modules"

f = os.path.join(base, "Применение KKS - ОП СПб.xlsx")
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb['Применение KKS']

# Read all rows with data
print("=== Применение KKS (full) ===")
for r in range(1, min(ws.max_row + 1, 200)):
    vals = []
    for col in range(1, 6):
        v = ws.cell(row=r, column=col).value
        if v is not None:
            vals.append((openpyxl.utils.get_column_letter(col), v))
    if vals:
        print(f"Row {r}: {vals}")

wb.close()
