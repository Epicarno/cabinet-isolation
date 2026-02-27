# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r"c:\Users\Администратор\Desktop\Modules"

# SHKZIAV ТИ - full dump
f = os.path.join(base, "ИО ШКЗиАВ 03.1_09092025.xlsm")
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb['ТИ']
print("=== SHKZIAV: ТИ (all rows) ===")
for r in range(1, 30):
    a = ws.cell(row=r, column=1).value  # Наименование
    b = ws.cell(row=r, column=2).value  # Массив ТИ
    h = ws.cell(row=r, column=8).value  # Система
    i_col = ws.cell(row=r, column=9).value  # Обознач.
    if a is not None:
        print(f"  Row {r}: Name={a} | Array={b} | Sys={h} | Mark={i_col}")
wb.close()

# Also look at ТС sheets for all 3
for name, fname in [
    ("SHD", "SHD_03_1_10092025 v1_01.xlsm"),
    ("SHUOD", "ИО ШУОД 03.1_18112025.xlsm"),
    ("SHKZIAV", "ИО ШКЗиАВ 03.1_09092025.xlsm"),
]:
    f = os.path.join(base, fname)
    wb = openpyxl.load_workbook(f, data_only=True)
    
    # Get Таблица сигналов to understand KKS naming
    ws = wb['Таблица сигналов']
    print(f"\n=== {name}: Таблица сигналов (first 10 rows) ===")
    for r in range(1, 11):
        vals = []
        for col in range(1, 20):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                vals.append((openpyxl.utils.get_column_letter(col), v))
        if vals:
            print(f"  Row {r}: {vals}")
    wb.close()
