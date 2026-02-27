# -*- coding: utf-8 -*-
import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

base = r"c:\Users\Администратор\Desktop\Modules"

files_info = [
    ("SHD", "SHD_03_1_10092025 v1_01.xlsm"),
    ("SHUOD", "ИО ШУОД 03.1_18112025.xlsm"),
    ("SHKZIAV", "ИО ШКЗиАВ 03.1_09092025.xlsm"),
]

for name, fname in files_info:
    fpath = os.path.join(base, fname)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    
    # LIST5
    ws = wb['LIST5']
    print(f"\n{'='*80}")
    print(f"=== {name}: LIST5 (физические сигналы) ===")
    for row_idx in range(1, 120):
        a = ws.cell(row=row_idx, column=1).value  # №
        b = ws.cell(row=row_idx, column=2).value  # Наименование
        c = ws.cell(row=row_idx, column=3).value  # Имя переменной
        if b is not None and 'Резерв' not in str(b):
            print(f"  Row {row_idx}: №={a} | Name={b} | Var={c}")
    
    # Check TI sheet for SHKZIAV
    if name == "SHKZIAV":
        ws_ti = wb['ТИ']
        print(f"\n--- {name}: лист ТИ ---")
        for row_idx in range(1, 100):
            vals = []
            for col in range(1, 15):
                v = ws_ti.cell(row=row_idx, column=col).value
                if v is not None:
                    vals.append((openpyxl.utils.get_column_letter(col), v))
            if vals:
                # Look for "oip" related content
                row_text = ' '.join(str(v[1]) for v in vals).lower()
                if 'oip' in row_text or 'оip' in row_text or row_idx <= 5:
                    print(f"  Row {row_idx}: {vals}")
    
    wb.close()
