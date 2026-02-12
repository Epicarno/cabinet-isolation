#!/usr/bin/env python3
"""
rename_kks.py — Переименование точек данных (datapoints) по стандарту KKS.

Читает правила из Excel-файла «Применение KKS - ОП СПб.xlsx»:
  Лист 1  — примеры преобразования OLD DP → NEW KKS (включая исключения)
  Лист 2  — таблица соответствия «шкаф → блок» (SHD_03_1 → B3)

Обрабатывает:
  1) XML мнемосхемы  (LCSMnemo/*/  )  — атрибут Name="" в <reference>
  2) XML объекты     (objects_*/    )  — при наличии DP в Name=""

Правила:
  ОСНОВНОЕ:  старый DP начинается с <ШКАФ>…  →  добавляем «<БЛОК>» спереди
             Например: SHD_03_1>P3_V3>DI>KZ1 → B3>SHD_03_1>P3_V3>DI>KZ1

  ИСКЛЮЧЕНИЯ (RESTRUCTURED): некоторые DP перестраивают сегменты помимо
             добавления блока.  Они берутся напрямую из колонок B→C листа 1.

Режимы:
  --dry-run  (по умолчанию) — только отчёт, файлы не трогаем
  --apply                  — реально записываем изменения

Использование:
  python rename_kks.py                   # dry-run по всем шкафам
  python rename_kks.py --apply           # применить замены
  python rename_kks.py --dry-run SHD_12  # только шкаф SHD_12
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from collections import defaultdict

# ---------------------------------------------------------------------------
# Пути
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Маппинг физических точек (StSign / OIP) → подсистема
# Формат KKS: Block>Cabinet>Subsystem>SignalType>PointName
# ---------------------------------------------------------------------------
# Обратный индекс строится автоматически функцией _build_physical_index()

PHYSICAL_GROUPS: dict[str, dict[str, dict[str, list[int]]]] = {
    # ---- SHUOD_03_1: организовано по помещениям (двери) ----
    "SHUOD_03_1": {
        "D109":        {"DI": [1, 2, 3, 4, 65, 69]},
        "D121A":       {"DI": [5, 6, 7, 8, 66, 70]},
        "D121":        {"DI": [9, 10, 11, 12, 67, 71]},
        "D120":        {"DI": [13, 14, 15, 16, 68, 72]},
        "D119":        {"DI": [17, 18, 19, 20, 73, 77]},
        "FIRE":        {"DI": [21]},
        "DIAG":        {"DI": [27, 28, 29, 30, 31, 32, 62, 63, 64, 97]},
        "RSRV_DI_A14": {"DI": list(range(22, 27))},
        "RSRV_DI_A15": {"DI": list(range(33, 62))},
        "RSRV_DO_A13": {"DO": [74, 75, 76, 78, 79, 80]},
        "RSRV_DO_A14": {"DO": list(range(81, 97))},
    },
    # ---- SHKZIAV_03_1: организовано по функциям ----
    "SHKZIAV_03_1": {
        "ALRT_ST":   {"DI": [1, 17, 18, 19]},
        "ALRT_R106": {"DI": [2, 20, 21, 22]},
        "GAS":       {"DI": [3, 4], "AI": [1]},       # OIP_1 = AI
        "EVENT":     {"DI": [23, 24]},
        "FIRE":      {"DI": [5]},
        "DIAG":      {"DI": [8, 9, 10, 11, 12, 13, 14, 15, 16, 33]},
        "RSRV_DI":   {"DI": [6, 7]},
        "RSRV_DO":   {"DO": list(range(25, 33))},
        "RSRV_AI":   {"AI": list(range(2, 13))},
    },
}

# Алиасы шкафов: вариант написания в XML → каноническое имя в PHYSICAL_GROUPS
# (в XML может быть SHKZIAV_03_1, а в Excel/DPL — SHKZiAV_03_1)
_CAB_ALIASES: dict[str, str] = {
    "SHKZIAV_03_1": "SHKZIAV_03_1",
    "SHKZiAV_03_1": "SHKZIAV_03_1",
}

# Обратный индекс: (cabinet, point_type, number) → (subsystem, signal_type)
# point_type = "StSign" или "OIP"
# cabinet — каноническое имя из PHYSICAL_GROUPS
_PHYSICAL_INDEX: dict[tuple[str, str, int], tuple[str, str]] = {}


def _build_physical_index() -> None:
    """Строим обратный индекс из PHYSICAL_GROUPS."""
    for cab, groups in PHYSICAL_GROUPS.items():
        for subsystem, signals in groups.items():
            for sig_type, numbers in signals.items():
                for n in numbers:
                    # DI/DO → StSign, AI → OIP
                    pt = "OIP" if sig_type == "AI" else "StSign"
                    _PHYSICAL_INDEX[(cab, pt, n)] = (subsystem, sig_type)


def _normalize_cab(cab: str) -> str:
    """Нормализуем имя шкафа: SHKZiAV_03_1 → SHKZIAV_03_1 и т.д."""
    return _CAB_ALIASES.get(cab, cab)


_build_physical_index()
MODULES_DIR = SCRIPT_DIR.parent
EXCEL_FILE  = MODULES_DIR / "Применение KKS - ОП СПб.xlsx"
REPORT_DIR  = MODULES_DIR / "reports"
MNEMO_DIR   = MODULES_DIR / "ventcontent" / "panels" / "vision" / "LCSMnemo"


def load_excel(xlsx: Path) -> tuple[dict[str, str], dict[str, str]]:
    """Загружаем правила из Excel.

    Returns:
        cab_to_block  — {cabinet_name: block_prefix}
        explicit_map  — {old_dp: new_dp}  для RESTRUCTURED и прочих
                        (включая записи без разделителей ">")
    """
    try:
        import openpyxl
    except ImportError:
        print("ОШИБКА: установите openpyxl  (pip install openpyxl)", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    # --- Лист 2: шкаф → блок ---
    ws2 = wb[wb.sheetnames[1]]
    cab_to_block: dict[str, str] = {}
    for r in range(2, ws2.max_row + 1):
        c_val = ws2.cell(r, 3).value
        if not c_val or ">" not in str(c_val):
            continue
        parts = str(c_val).rstrip(">").split(">")
        if len(parts) >= 2 and parts[1]:
            cab_to_block[parts[1]] = parts[0]  # SHD_03_1 → B3

    # --- Лист 1: явные маппинги (все + выявляем RESTRUCTURED) ---
    ws1 = wb[wb.sheetnames[0]]
    explicit_map: dict[str, str] = {}
    for r in range(2, ws1.max_row + 1):
        b = ws1.cell(r, 2).value
        c = ws1.cell(r, 3).value
        e = ws1.cell(r, 5).value  # альтернативный KKS
        if not b or not c:
            continue
        old = str(b).strip()
        new = str(c).strip()
        if "KKS" in new:
            continue  # заголовок

        # Проверяем — простой ли это PREFIX случай
        if ">" in old:
            old_parts = old.split(">")
            new_parts = new.split(">")
            if len(new_parts) == len(old_parts) + 1 and new_parts[1:] == old_parts:
                continue  # обычный PREFIX, обработается общим правилом
        
        # Всё остальное — явный маппинг
        explicit_map[old] = new

    wb.close()
    return cab_to_block, explicit_map


def extract_cabinet(dp_name: str, cab_to_block: dict[str, str]) -> str | None:
    """Извлекаем имя шкафа из DP.

    DP может быть:
      SHD_03_1>P3_V3>DI>KZ1   — первый сегмент = шкаф
      SHD_7_A1_2               — без >, весь DP начинается с имени шкафа
      SHKZiAV_03_1_OIP_1       — без >, prefix matching по cab_to_block
    """
    if ">" in dp_name:
        cab_candidate = dp_name.split(">")[0]
        found = _find_cab(cab_candidate, cab_to_block)
        if found:
            return found
    else:
        # Без разделителя — ищем самый длинный подходящий шкаф-префикс
        best = None
        for cab in cab_to_block:
            if dp_name.lower().startswith(cab.lower()) and (best is None or len(cab) > len(best)):
                best = cab
        return best
    return None


def transform_dp(dp_name: str,
                 cab_to_block: dict[str, str],
                 explicit_map: dict[str, str]) -> str | None:
    """Преобразуем DP по правилам KKS.

    Returns:
        Новое имя или None если преобразование не нужно / невозможно.
    """
    # 1) Если DP уже начинается с блока — пропускаем
    for block in set(cab_to_block.values()):
        if dp_name.startswith(block + ">"):
            return None  # уже в формате KKS

    # 2) Явный маппинг (RESTRUCTURED + без разделителей)
    if dp_name in explicit_map:
        return explicit_map[dp_name]

    # 3) Физические точки: Cabinet_StSign_N / Cabinet_OIP_N
    #    → Block>Cabinet>Subsystem>SignalType>PointName
    phys = _try_physical_transform(dp_name, cab_to_block)
    if phys is not None:
        return phys

    # 4) Общее правило: добавить БЛОК> перед DP
    cab = extract_cabinet(dp_name, cab_to_block)
    if cab is None:
        return None  # неизвестный шкаф — не трогаем
    block = cab_to_block[cab]
    return f"{block}>{dp_name}"


# Regex для разбора физических точек: Cabinet_StSign_N или Cabinet_OIP_N
_PHYS_RE = re.compile(r'^(.+?)_(StSign|OIP)_(\d+)$')


def _find_cab(name: str, cab_to_block: dict[str, str]) -> str | None:
    """Ищем шкаф в cab_to_block, допуская разницу в регистре (SHKZiAV vs SHKZIAV)."""
    if name in cab_to_block:
        return name
    low = name.lower()
    for cab in cab_to_block:
        if cab.lower() == low:
            return cab
    return None


def _try_physical_transform(dp_name: str,
                            cab_to_block: dict[str, str]) -> str | None:
    """Пытаемся преобразовать физическую точку через PHYSICAL_GROUPS.

    Формат входа:  SHUOD_03_1_StSign_3
    Формат выхода: B3>SHUOD_03_1>D109>DI>StSign_3
    """
    m = _PHYS_RE.match(dp_name)
    if not m:
        return None

    raw_cab = m.group(1)   # SHUOD_03_1  или  SHKZiAV_03_1 / SHKZIAV_03_1
    pt      = m.group(2)   # StSign / OIP
    num     = int(m.group(3))  # 3

    # Ищем шкаф в cab_to_block (case-insensitive)
    real_cab = _find_cab(raw_cab, cab_to_block)
    if real_cab is None:
        return None

    canon_cab = _normalize_cab(raw_cab)
    key = (canon_cab, pt, num)
    if key not in _PHYSICAL_INDEX:
        # Точка не в маппинге — используем общее правило (Block>flatDP)
        block = cab_to_block[real_cab]
        return f"{block}>{dp_name}"

    subsystem, sig_type = _PHYSICAL_INDEX[key]
    block = cab_to_block[real_cab]
    # Сохраняем оригинальное написание шкафа из DP (raw_cab)
    return f"{block}>{raw_cab}>{subsystem}>{sig_type}>{pt}_{num}"


# ---------------------------------------------------------------------------
# Regex для поиска DP-подобных Name="" в XML
# ---------------------------------------------------------------------------
# Ловим Name="<ШКАФ>..." внутри тегов <reference> и <shape>
# Шкафы начинаются с SHD_, SHK, SHUOD_, SHUVO_, VVSBiS_ и т.д.
DP_NAME_RE = re.compile(
    r'(Name=")([A-Za-z][A-Za-z0-9_]*(?:>[A-Za-z0-9_>.\-]+)*?)(")'
)


def process_file(xml_path: Path,
                 cab_to_block: dict[str, str],
                 explicit_map: dict[str, str],
                 apply: bool) -> list[tuple[str, str]]:
    """Обрабатываем один XML-файл.

    Returns:
        Список (old_dp, new_dp) выполненных замен.
    """
    try:
        content = xml_path.read_text(encoding="utf-8")
    except Exception:
        return []

    changes: list[tuple[str, str]] = []

    def replacer(m: re.Match) -> str:
        prefix = m.group(1)   # Name="
        dp = m.group(2)       # SHD_03_1>P3_V3>DI>KZ1
        suffix = m.group(3)   # "

        new_dp = transform_dp(dp, cab_to_block, explicit_map)
        if new_dp is None:
            return m.group(0)  # без изменений

        changes.append((dp, new_dp))
        return f"{prefix}{new_dp}{suffix}"

    new_content = DP_NAME_RE.sub(replacer, content)

    if changes and apply:
        xml_path.write_text(new_content, encoding="utf-8", newline="\n")

    return changes


def collect_xml_files(cab_to_block: dict[str, str],
                      cabinets_filter: list[str] | None) -> list[Path]:
    """Собираем все XML-файлы для обработки: мнемосхемы + объекты."""
    xmls: list[Path] = []

    # Мнемосхемы
    if MNEMO_DIR.is_dir():
        for cab_dir in sorted(MNEMO_DIR.iterdir()):
            if not cab_dir.is_dir():
                continue
            if cabinets_filter and cab_dir.name not in cabinets_filter:
                continue
            xmls.extend(sorted(cab_dir.rglob("*.xml")))

    # Объекты (objects_*)
    for obj_dir in sorted(MODULES_DIR.glob("objects_*")):
        if not obj_dir.is_dir():
            continue
        cab_name = obj_dir.name.removeprefix("objects_")
        if cabinets_filter and cab_name not in cabinets_filter:
            continue
        xmls.extend(sorted(obj_dir.rglob("*.xml")))

    # Также обработаем output/ если есть
    output_dir = MODULES_DIR / "output"
    if output_dir.is_dir():
        for cab_dir in sorted(output_dir.iterdir()):
            if not cab_dir.is_dir():
                continue
            if cabinets_filter and cab_dir.name not in cabinets_filter:
                continue
            xmls.extend(sorted(cab_dir.rglob("*.xml")))

    return xmls


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Переименование точек данных по стандарту KKS"
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Реально записать изменения (по умолчанию — dry-run)"
    )
    parser.add_argument(
        "--dry-run", action="store_true", default=True,
        help="Только отчёт, файлы не трогаем (по умолчанию)"
    )
    parser.add_argument(
        "cabinets", nargs="*",
        help="Шкафы для обработки (по умолчанию — все)"
    )
    args = parser.parse_args()

    if args.apply:
        args.dry_run = False

    if not EXCEL_FILE.exists():
        print(f"ОШИБКА: не найден файл {EXCEL_FILE}", file=sys.stderr)
        sys.exit(1)

    # Загружаем правила
    print(f"Загрузка правил из {EXCEL_FILE.name} ...")
    cab_to_block, explicit_map = load_excel(EXCEL_FILE)
    print(f"  Шкафов в таблице: {len(cab_to_block)}")
    print(f"  Явных маппингов (исключения): {len(explicit_map)}")

    # Фильтр шкафов
    cabinets_filter = args.cabinets if args.cabinets else None

    # Собираем файлы
    xmls = collect_xml_files(cab_to_block, cabinets_filter)
    print(f"  XML-файлов для обработки: {len(xmls)}")

    if not xmls:
        print("Нет файлов для обработки.")
        return

    # Обрабатываем
    mode_label = "APPLY" if args.apply else "DRY-RUN"
    print(f"\nРежим: {mode_label}\n")

    total_changes = 0
    files_changed = 0
    all_changes: list[tuple[Path, str, str]] = []

    for xml in xmls:
        changes = process_file(xml, cab_to_block, explicit_map, apply=args.apply)
        if changes:
            files_changed += 1
            total_changes += len(changes)
            rel = xml.relative_to(MODULES_DIR)
            for old, new in changes:
                all_changes.append((rel, old, new))

    # Статистика
    print(f"{'='*60}")
    print(f"Итого: {total_changes} замен в {files_changed} файлах")
    print(f"{'='*60}")

    # Группируем по типу замены
    change_types: dict[str, int] = defaultdict(int)
    for _, old, new in all_changes:
        if _PHYS_RE.match(old):
            change_types["PHYSICAL (StSign/OIP → подсистема)"] += 1
        elif ">" not in old:
            change_types["flat DP (без >)"] += 1
        elif new.split(">")[1:] == old.split(">"):
            change_types["PREFIX (добавлен блок)"] += 1
        else:
            change_types["RESTRUCTURED (перестройка)"] += 1

    if change_types:
        print("\nТипы замен:")
        for ctype, count in sorted(change_types.items(), key=lambda x: -x[1]):
            print(f"  {count:5d}  {ctype}")

    # Сохраняем отчёт
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_file = REPORT_DIR / "kks_rename.txt"
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(f"KKS Rename Report ({mode_label})\n")
        f.write(f"{'='*60}\n")
        f.write(f"Замен: {total_changes}, файлов: {files_changed}\n\n")

        # Группируем по файлу
        by_file: dict[Path, list[tuple[str, str]]] = defaultdict(list)
        for rel, old, new in all_changes:
            by_file[rel].append((old, new))

        for rel in sorted(by_file):
            f.write(f"\n--- {rel} ---\n")
            for old, new in by_file[rel]:
                f.write(f"  {old}\n")
                f.write(f"    => {new}\n")

    print(f"\nОтчёт: {report_file}")

    # Показываем несколько примеров
    if all_changes:
        print(f"\nПримеры замен (первые 10):")
        for rel, old, new in all_changes[:10]:
            print(f"  {rel}")
            print(f"    {old}  =>  {new}")


if __name__ == "__main__":
    main()
