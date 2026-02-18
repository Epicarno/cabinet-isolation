# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys, os

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(ws, title, data):
    ws.title = title
    headers = ["№ LIST5", "Старое имя (StSign)", "Описание", "Подсистема", "Новое имя KKS"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 45

# ============ SHD 03.1 ============
shd_data = [
    [1,  "SHD_03_1_StSign_1",  "Водомерный узел. Расходомер (B1-FGE1)",                      "VMU", "B3>SHD_03_1>VMU>DI>FGE1"],
    [6,  "SHD_03_1_StSign_6",  "ШД.03.1 - Исправность преобразователя интерфейсов A6",       "NAV", "B3>SHD_03_1>NAV>DI>A6"],
    [7,  "SHD_03_1_StSign_7",  "ШД.03.1 - Исправность коммутатора A3 (осн.)",                "NAV", "B3>SHD_03_1>NAV>DI>A3"],
    [8,  "SHD_03_1_StSign_8",  "ШД.03.1 - Исправность коммутатора A4 (рез.)",                "NAV", "B3>SHD_03_1>NAV>DI>A4"],
    [9,  "SHD_03_1_StSign_9",  "ШД.03.1 - Исправность ИБП 220В GU1",                        "NAV", "B3>SHD_03_1>NAV>DI>GU1"],
    [10, "SHD_03_1_StSign_10", "ШД.03.1 - Исправность оборудования питания 24В",             "NAV", "B3>SHD_03_1>NAV>DI>POWER"],
    [11, "SHD_03_1_StSign_11", "ШД.03.1 - Наличие напряжения на вводе №1",                   "NAV", "B3>SHD_03_1>NAV>DI>SUPPLY1"],
    [12, "SHD_03_1_StSign_12", "ШД.03.1 - Наличие напряжения на вводе №2",                   "NAV", "B3>SHD_03_1>NAV>DI>SUPPLY2"],
    [13, "SHD_03_1_StSign_13", "ШД.03.1 - Превышение температуры",                           "NAV", "B3>SHD_03_1>NAV>DI>OVERHEAT"],
    [14, "SHD_03_1_StSign_14", "ШД.03.1 - Включена вентиляция",                              "NAV", "B3>SHD_03_1>NAV>DI>FAN"],
    [15, "SHD_03_1_StSign_15", "ШД.03.1 - Двери шкафа открыты",                              "NAV", "B3>SHD_03_1>NAV>DI>DOOR"],
    [16, "SHD_03_1_StSign_16", "ШД.03.1 - Низкий уровень сопротивления изоляции 24В (UV1)",  "NAV", "B3>SHD_03_1>NAV>DI>UV1"],
    [17, "SHD_03_1_StSign_17", "ШД.03.1 - Наличие напряжения на вводе №1,2",                 "NAV", "B3>SHD_03_1>NAV>DI>SUPPLY12"],
]
ws1 = wb.active
write_sheet(ws1, "SHD_03_1", shd_data)

# ============ SHUOD 03.1 ============
shuod_data = [
    [1,  "SHUOD_03_1_StSign_1",  "Помещение 109 - Разблокировка доступа в помещение 118 (SB1.3)",                "SKUD_118",  "B3>SHUOD_03_1>SKUD_118>DI>SB1_3"],
    [2,  "SHUOD_03_1_StSign_2",  "Помещение 118 - Разблокировка доступа в помещение 118 (SB1.1)",                "SKUD_118",  "B3>SHUOD_03_1>SKUD_118>DI>SB1_1"],
    [3,  "SHUOD_03_1_StSign_3",  "Помещение 118 - Контроль положения дверей в помещение 118 (SQ1.1, SQ1.2)",     "SKUD_118",  "B3>SHUOD_03_1>SKUD_118>DI>SQ1"],
    [4,  "SHUOD_03_1_StSign_4",  "Помещение 118 - Аварийная разблокировка доступа в помещение 106 (SB1.2, HO)",  "SKUD_118",  "B3>SHUOD_03_1>SKUD_118>DI>SB1_2"],
    [5,  "SHUOD_03_1_StSign_5",  "Помещение 118 - Разблокировка доступа в помещение 121а (SB2.3)",               "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DI>SB2_3"],
    [6,  "SHUOD_03_1_StSign_6",  "Помещение 121а - Разблокировка доступа в помещение 118 (SB2.1)",               "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DI>SB2_1"],
    [7,  "SHUOD_03_1_StSign_7",  "Помещение 121a - Контроль положения дверей в помещение 121а (SQ2)",            "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DI>SQ2"],
    [8,  "SHUOD_03_1_StSign_8",  "Помещение 121а - Аварийная разблокировка доступа в помещение 118 (SB2.2, HO)", "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DI>SB2_2"],
    [9,  "SHUOD_03_1_StSign_9",  "Помещение 118 - Разблокировка доступа в помещение 121 (SB3.3)",                "SKUD_121",  "B3>SHUOD_03_1>SKUD_121>DI>SB3_3"],
    [10, "SHUOD_03_1_StSign_10", "Помещение 121 - Разблокировка доступа в помещение 118 (SB3.1)",                "SKUD_121",  "B3>SHUOD_03_1>SKUD_121>DI>SB3_1"],
    [11, "SHUOD_03_1_StSign_11", "Помещение 121 - Контроль положения дверей в помещение 121 (SQ3.1, SQ3.2)",     "SKUD_121",  "B3>SHUOD_03_1>SKUD_121>DI>SQ3"],
    [12, "SHUOD_03_1_StSign_12", "Помещение 121 - Аварийная разблокировка доступа в помещение 118 (SB3.2, HO)",  "SKUD_121",  "B3>SHUOD_03_1>SKUD_121>DI>SB3_2"],
    [13, "SHUOD_03_1_StSign_13", "Помещение 118 - Разблокировка доступа в помещение 120 (SB4.3)",                "SKUD_120",  "B3>SHUOD_03_1>SKUD_120>DI>SB4_3"],
    [14, "SHUOD_03_1_StSign_14", "Помещение 120 - Разблокировка доступа в помещение 118 (SB4.1)",                "SKUD_120",  "B3>SHUOD_03_1>SKUD_120>DI>SB4_1"],
    [15, "SHUOD_03_1_StSign_15", "Помещение 120 - Контроль положения дверей в помещение 120 (SQ4.1, SQ4.2)",     "SKUD_120",  "B3>SHUOD_03_1>SKUD_120>DI>SQ4"],
    [16, "SHUOD_03_1_StSign_16", "Помещение 120 - Аварийная разблокировка доступа в помещение 118 (SB4.2, HO)",  "SKUD_120",  "B3>SHUOD_03_1>SKUD_120>DI>SB4_2"],
    [17, "SHUOD_03_1_StSign_17", "Помещение 118 - Разблокировка доступа в помещение 119 (SB5.3)",                "SKUD_119",  "B3>SHUOD_03_1>SKUD_119>DI>SB5_3"],
    [18, "SHUOD_03_1_StSign_18", "Помещение 119 - Разблокировка доступа в помещение 118 (SB5.1)",                "SKUD_119",  "B3>SHUOD_03_1>SKUD_119>DI>SB5_1"],
    [19, "SHUOD_03_1_StSign_19", "Помещение 119 - Контроль положения дверей в помещение 119 (SQ5.1, SQ5.2)",     "SKUD_119",  "B3>SHUOD_03_1>SKUD_119>DI>SQ5"],
    [20, "SHUOD_03_1_StSign_20", "Помещение 119 - Аварийная разблокировка доступа в помещение 118 (SB5.2, HO)",  "SKUD_119",  "B3>SHUOD_03_1>SKUD_119>DI>SB5_2"],
    [21, "SHUOD_03_1_StSign_21", "Сигнал \"ПОЖАР\"",                                                             "FIRE",      "B3>SHUOD_03_1>FIRE>DI>FIRE"],
    [27, "SHUOD_03_1_StSign_27", "ШУОД 03.1 - Наличие напряжения на вводе №1",                                  "NAV",       "B3>SHUOD_03_1>NAV>DI>SUPPLY1"],
    [28, "SHUOD_03_1_StSign_28", "ШУОД 03.1 - Наличие напряжения на вводе №2",                                  "NAV",       "B3>SHUOD_03_1>NAV>DI>SUPPLY2"],
    [29, "SHUOD_03_1_StSign_29", "ШУОД 03.1 - Превышение температуры",                                          "NAV",       "B3>SHUOD_03_1>NAV>DI>OVERHEAT"],
    [30, "SHUOD_03_1_StSign_30", "ШУОД 03.1 - Включена вентиляция",                                             "NAV",       "B3>SHUOD_03_1>NAV>DI>FAN"],
    [31, "SHUOD_03_1_StSign_31", "ШУОД 03.1 - Двери шкафа открыты",                                             "NAV",       "B3>SHUOD_03_1>NAV>DI>DOOR"],
    [32, "SHUOD_03_1_StSign_32", "ШУОД 03.1 - Низкий уровень сопротивления изоляции 24В (UV1)",                 "NAV",       "B3>SHUOD_03_1>NAV>DI>UV1"],
    [62, "SHUOD_03_1_StSign_62", "ШУОД 03.1 - Исправность оборудования питания 24В",                            "NAV",       "B3>SHUOD_03_1>NAV>DI>POWER"],
    [63, "SHUOD_03_1_StSign_63", "ШУОД 03.1 - Исправность коммутатора A4",                                      "NAV",       "B3>SHUOD_03_1>NAV>DI>A4"],
    [64, "SHUOD_03_1_StSign_64", "ШУОД 03.1 - Исправность ИБП 220В GU1",                                       "NAV",       "B3>SHUOD_03_1>NAV>DI>GU1"],
    [65, "SHUOD_03_1_StSign_65", "Помещение 118 - Разблокировка доступа в помещение 109 (SB1.2, НЗ)",           "SKUD_118",  "B3>SHUOD_03_1>SKUD_118>DO>SB1_2"],
    [66, "SHUOD_03_1_StSign_66", "Помещение 121а - Разблокировка доступа в помещение 118 (SB2.2, НЗ)",          "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DO>SB2_2"],
    [67, "SHUOD_03_1_StSign_67", "Помещение 121 - Разблокировка доступа в помещение 118 (SB3.2, НЗ)",           "SKUD_121",  "B3>SHUOD_03_1>SKUD_121>DO>SB3_2"],
    [68, "SHUOD_03_1_StSign_68", "Помещение 120 - Разблокировка доступа в помещение 118 (SB4.2, НЗ)",           "SKUD_120",  "B3>SHUOD_03_1>SKUD_120>DO>SB4_2"],
    [69, "SHUOD_03_1_StSign_69", "Помещение 118 - Звуковое оповещение о нарушении блокировки помещения 118 (HA1)", "SKUD_118", "B3>SHUOD_03_1>SKUD_118>DO>HA1"],
    [70, "SHUOD_03_1_StSign_70", "Помещение 121а - Звуковое оповещение о нарушении блокировки помещения 121а (HA2)", "SKUD_121A", "B3>SHUOD_03_1>SKUD_121A>DO>HA2"],
    [71, "SHUOD_03_1_StSign_71", "Помещение 121 - Звуковое оповещение о нарушении блокировки помещения 121 (HA3)",  "SKUD_121", "B3>SHUOD_03_1>SKUD_121>DO>HA3"],
    [72, "SHUOD_03_1_StSign_72", "Помещение 120 - Звуковое оповещение о нарушении блокировки помещения 120 (HA4)",  "SKUD_120", "B3>SHUOD_03_1>SKUD_120>DO>HA4"],
    [73, "SHUOD_03_1_StSign_73", "Помещение 119 - Разблокировка доступа в помещение 118 (SB5.2, НЗ)",           "SKUD_119",  "B3>SHUOD_03_1>SKUD_119>DO>SB5_2"],
    [77, "SHUOD_03_1_StSign_77", "Помещение 119 - Звуковое оповещение о нарушении блокировки помещения 119 (HA5)", "SKUD_119", "B3>SHUOD_03_1>SKUD_119>DO>HA5"],
    [97, "SHUOD_03_1_StSign_97", "ШУОД 03.1 - Наличие напряжения на вводе №1,2",                                "NAV",       "B3>SHUOD_03_1>NAV>DI>SUPPLY12"],
]
ws2 = wb.create_sheet()
write_sheet(ws2, "SHUOD_03_1", shuod_data)

# ============ SHKZIAV 03.1 ============
shkziav_data = [
    [1,  "SHKZiAV_03_1_StSign_1",  "Улица (у двери в помещение 107). Квитирование оповещение (3-SB1)",                   "GAS",  "B3>SHKZiAV_03_1>GAS>DI>3_SB1"],
    [2,  "SHKZiAV_03_1_StSign_2",  "Улица (у двери в помещение 106). Квитирование оповещение (3-SB2)",                   "GAS",  "B3>SHKZiAV_03_1>GAS>DI>3_SB2"],
    [3,  "SHKZiAV_03_1_StSign_3",  "Помещение 107. Датчик загазованности (3-A1) >10% НКПР",                              "GAS",  "B3>SHKZiAV_03_1>GAS>DI>3_A1"],
    [4,  "SHKZiAV_03_1_StSign_4",  "Помещение 107. Датчик загазованности (3-A1) неисправность",                          "GAS",  "B3>SHKZiAV_03_1>GAS>DI>3_A1_FAULT"],
    [5,  "SHKZiAV_03_1_StSign_5",  "Сигнал \"ПОЖАР\"",                                                                   "FIRE", "B3>SHKZiAV_03_1>FIRE>DI>FIRE"],
    [8,  "SHKZiAV_03_1_StSign_8",  "ШКЗиАВ.03.1 - Исправность оборудования питания 24В",                                "NAV",  "B3>SHKZiAV_03_1>NAV>DI>POWER"],
    [9,  "SHKZiAV_03_1_StSign_9",  "ШКЗиАВ.03.1 - Исправность коммутатора A4 (рез.)",                                   "NAV",  "B3>SHKZiAV_03_1>NAV>DI>A4"],
    [10, "SHKZiAV_03_1_StSign_10", "ШКЗиАВ.03.1 - Исправность ИБП 220В GU1",                                            "NAV",  "B3>SHKZiAV_03_1>NAV>DI>GU1"],
    [11, "SHKZiAV_03_1_StSign_11", "ШКЗиАВ.03.1 - Наличие напряжения на вводе №1",                                      "NAV",  "B3>SHKZiAV_03_1>NAV>DI>SUPPLY1"],
    [12, "SHKZiAV_03_1_StSign_12", "ШКЗиАВ.03.1 - Наличие напряжения на вводе №2",                                      "NAV",  "B3>SHKZiAV_03_1>NAV>DI>SUPPLY2"],
    [13, "SHKZiAV_03_1_StSign_13", "ШКЗиАВ.03.1 - Превышение температуры",                                              "NAV",  "B3>SHKZiAV_03_1>NAV>DI>OVERHEAT"],
    [14, "SHKZiAV_03_1_StSign_14", "ШКЗиАВ.03.1 - Включена вентиляция",                                                 "NAV",  "B3>SHKZiAV_03_1>NAV>DI>FAN"],
    [15, "SHKZiAV_03_1_StSign_15", "ШКЗиАВ.03.1 - Двери шкафа открыты",                                                 "NAV",  "B3>SHKZiAV_03_1>NAV>DI>DOOR"],
    [16, "SHKZiAV_03_1_StSign_16", "ШКЗиАВ.03.1 - Низкий уровень сопротивления изоляции 24В (UV1)",                     "NAV",  "B3>SHKZiAV_03_1>NAV>DI>UV1"],
    [17, "SHKZiAV_03_1_StSign_17", "Улица (у двери в помещение 107) Светозвуковое оповещение (3-HLA1) Свет",             "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HLA1_L"],
    [18, "SHKZiAV_03_1_StSign_18", "Улица (у двери в помещение 107) Светозвуковое оповещение (3-HLA1) Звук",             "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HLA1_S"],
    [19, "SHKZiAV_03_1_StSign_19", "Улица (у двери в помещение 107) Табло загазованности (3-HL1)",                        "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HL1"],
    [20, "SHKZiAV_03_1_StSign_20", "Помещение 106 (у двери в помещение 107) Светозвуковое оповещение (3-HLA2) Свет",     "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HLA2_L"],
    [21, "SHKZiAV_03_1_StSign_21", "Помещение 106 (у двери в помещение 107) Светозвуковое оповещение (3-HLA2) Звук",     "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HLA2_S"],
    [22, "SHKZiAV_03_1_StSign_22", "Помещение 106 (у двери в помещение 107) Табло загазованности (3-HL2)",                "GAS",  "B3>SHKZiAV_03_1>GAS>DO>3_HL2"],
    [23, "SHKZiAV_03_1_StSign_23", "Отключение вытяжки (поз. 63)",                                                       "GAS",  "B3>SHKZiAV_03_1>GAS>DO>EXHAUST_OFF"],
    [24, "SHKZiAV_03_1_StSign_24", "Включение аварийной вентиляции (ЯУВ8)",                                              "GAS",  "B3>SHKZiAV_03_1>GAS>DO>EMRG_VENT"],
    [33, "SHKZiAV_03_1_StSign_33", "ШКЗиАВ.03.1 - Наличие напряжения на вводе №1,2",                                    "NAV",  "B3>SHKZiAV_03_1>NAV>DI>SUPPLY12"],
]
ws3 = wb.create_sheet()
write_sheet(ws3, "SHKZiAV_03_1", shkziav_data)

# ============ OIP (ТИ) ============
ws4 = wb.create_sheet("OIP")
headers = ["№ ТИ", "Старое имя", "Описание", "Подсистема", "Новое имя KKS"]
for col, h in enumerate(headers, 1):
    c = ws4.cell(row=1, column=col, value=h)
    c.font = header_font_white
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = thin_border
oip_data = [1, "SHKZiAV_03_1_OIP_1", "Помещение 107. Датчик загазованности (3-A1). Текущая концентрация", "GAS", "B3>SHKZiAV_03_1>OIP_1"]
for col_idx, val in enumerate(oip_data, 1):
    c = ws4.cell(row=2, column=col_idx, value=val)
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True)
ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 65
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['E'].width = 45

out = os.path.join(r"c:\Users\Администратор\Desktop\Modules\reports", "kks_stsign_03_1.xlsx")
wb.save(out)
print(f"Saved: {out}")
